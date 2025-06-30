import streamlit as st
import pandas as pd
import numpy as np
import json
from datetime import datetime
from openai import OpenAI
import matplotlib.pyplot as plt
from PIL import Image
import ast
import seaborn as sns
import openpyxl
from openpyxl import Workbook, load_workbook
import os
import os
import matplotlib.cm as cm
import matplotlib.colors as mcolors

from openai import AzureOpenAI

client = AzureOpenAI(
    api_key="51a34d40daef4d71a08b4d0707436b3d",  # Replace with env var in production
    api_version="2024-06-01",
    azure_endpoint="https://api.umgpt.umich.edu/azure-openai-api",
    organization="018827"
)

EXCEL_FILE = "bluechoice_data.xlsx"

def read_excel(sheet_name):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        wb.save(EXCEL_FILE)
    wb = load_workbook(EXCEL_FILE)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
        wb.save(EXCEL_FILE)
    ws = wb[sheet_name]
    return ws, wb

def write_excel(sheet_name, data, mode='append'):
    ws, wb = read_excel(sheet_name)
    if mode == 'overwrite':
        ws.delete_rows(1, ws.max_row)
    for row in data:
        ws.append(row)
    wb.save(EXCEL_FILE)

# ✅ FIRST Streamlit command
st.set_page_config(layout="wide")

# === Config and setup continues below ===
WEC_DESIGNS = ["Point Absorber", "Oscillating Water Column", "Oscillating Surge Flap"]

client = OpenAI(api_key=st.secrets["OPENAI_API_KEY"])

@st.cache_data(ttl=600)
def load_themes_and_subcriteria(sheet=None):
    ws, _ = read_excel("Themes Tab")
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data = rows[1:]
    theme_map = {}
    for row in data:
        theme, sub = row
        if theme and sub:
            theme_map.setdefault(theme.strip(), []).append(sub.strip())
    return list(theme_map.keys()), theme_map

def connect_to_google_sheets():
    return EXCEL_FILE

def create_themes_tab_with_subcriteria(sheet=None):
    ws, wb = read_excel("Themes Tab")
    if ws.max_row > 1:
        return  # Already initialized
    ws.append(["Theme", "Subcriterion"])
    theme_data = [
        ("Functional Efficiency", "Electricity reliability and security"),
        ("Functional Efficiency", "Electricity affordability"),
        ("Environmental Sustainability", "Habitat of marine animals"),
        ("Environmental Sustainability", "Birds"),
        ("Environmental Sustainability", "Health"),
        ("Environmental Sustainability", "Noise"),
        ("Sense of Place", "Personal identity / connection to place"),
        ("Sense of Place", "Ocean / landscape view"),
        ("Community Prosperity", "Community benefits"),
        ("Community Prosperity", "Job opportunities"),
        ("Community Prosperity", "Tax revenues"),
        ("Community Prosperity", "Indirect economic effects"),
        ("Community Prosperity", "Tourism"),
        ("Marine Space Utilization", "Recreational fishing"),
        ("Marine Space Utilization", "Recreational boating")
    ]
    for theme, sub in theme_data:
        ws.append([theme, sub])
    wb.save(EXCEL_FILE)


sheet = connect_to_google_sheets()
create_themes_tab_with_subcriteria(sheet)  # Make sure the sheet exists
THEMES, THEME_SUBCRITERIA = load_themes_and_subcriteria(sheet)

def create_sheet_if_missing(sheet_path, title):
    """Creates a new sheet in the local Excel file if it doesn't exist."""
    if not os.path.exists(sheet_path):
        # Create workbook and add the requested sheet
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet
        wb.create_sheet(title)
        wb.save(sheet_path)
        return

    wb = openpyxl.load_workbook(sheet_path)
    if title not in wb.sheetnames:
        wb.create_sheet(title)
        wb.save(sheet_path)


def ensure_headers_if_missing(title):
    ws, wb = read_excel(title)  # Load worksheet and workbook

    existing_rows = list(ws.iter_rows(values_only=True))

    if not existing_rows or all(all(cell in [None, ""] for cell in row) for row in existing_rows):
        ws.delete_rows(1, ws.max_row)  # Clear existing content if any

        if title == "Community Feedback":
            headers = ["Timestamp", "Name", "Community"] + THEMES
        elif title == "Expert Tab":
            headers = ["Theme"] + WEC_DESIGNS
        elif title == "Final Rankings":
            headers = ["WEC Design", "Closeness to Ideal"]
        elif title == "Expert Contributions Tab":
            headers = ["Timestamp", "Name", "Theme", "Expertise Level"]
            for i in range(len(WEC_DESIGNS)):
                for j in range(i + 1, len(WEC_DESIGNS)):
                    headers.append(f"PCM_{WEC_DESIGNS[i]}_vs_{WEC_DESIGNS[j]}")
        else:
            headers = []

        if headers:
            ws.append(headers)
            wb.save(EXCEL_FILE)

def clear_all_google_sheet_tabs():
    """
    Connects to the Google Sheet and clears all worksheets (tabs) without deleting them.
    """
    try:
        sheet = connect_to_google_sheets()  # Uses your existing connection function
        for ws in sheet.worksheets():
            ws.clear()
        st.success("✅ All sheets have been cleared successfully.")
    except Exception as e:
        st.error(f"❌ Failed to clear sheets: {e}")

# === Consistency Check ===
def calculate_consistency_index(pcm):
    n = len(pcm)
    eigvals, _ = np.linalg.eig(pcm)
    lam_max = np.max(np.real(eigvals))
    CI = (lam_max - n) / (n - 1) if n > 1 else 0

    # Saaty's RI values (Robles-Algarin 2017)
    RI_lookup = {
        1: 0.0,
        2: 0.0,
        3: 0.52,
        4: 0.89,
        5: 1.11,
        6: 1.25,
        7: 1.35,
        8: 1.40,
        9: 1.45,
        10: 1.49
    }

    RI = RI_lookup.get(n, 1.49)  # Default to 1.49 if n > 10
    CR = CI / RI if RI != 0 else 0
    return round(CR, 4), round(lam_max, 3), round(CI, 4)


def ahp_weights(pcm):
    geom_mean = np.prod(pcm, axis=1) ** (1 / len(pcm))
    weights = geom_mean / np.sum(geom_mean)
    return weights

def get_whfs_from_likert_and_text(likert_score, explanation_text):
    l = max(1, likert_score - 1)
    m = likert_score
    u = min(5, likert_score + 1)

    prompt = f"""
    You are converting community concern explanations into Weighted Hesitant Fuzzy Scores (WHFS).
    The respondent gave a Likert rating of {likert_score} (out of 5) for this theme.

    Please read the explanation and assign weights to the levels {l}, {m}, and {u} based on how much the explanation reflects each.

    Respond strictly in the format:
    {{"scores": [[{l}, w1], [{m}, w2], [{u}, w3]]}}
    Make sure weights sum to 1.

    Explanation:
    \"\"\"{explanation_text}\"\"\"
    """

    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1
        )
        import re
        raw = response.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = re.sub(r"^```(?:json)?", "", raw)
            raw = re.sub(r"```$", "", raw)
        parsed = json.loads(raw)
        # import pdb; pdb.set_trace()
        return parsed.get("scores", [])

    except Exception as e:
        return []  # fallback empty if GPT fails


# === Defuzzification ===
def defuzzify(tfn): return sum(tfn)/3

def aggregate_whfs_scores(df):
    theme_scores = {theme: [] for theme in THEMES}
    for i, row in df.iterrows():
        for theme in THEMES:
            raw = row.get(theme, "")
            try:
                score_obj = ast.literal_eval(raw) if isinstance(raw, str) else raw
                if isinstance(score_obj, list):  # WHFS is a list of TFNs
                    for tfn in score_obj:
                        if isinstance(tfn, list) and len(tfn) == 3:
                            theme_scores[theme].append(defuzzify(tfn))
            except Exception as e:
                print(f"Error parsing fuzzy score for row {i}, theme {theme}: {e}")
                continue
    weights = {t: np.mean(vals) for t, vals in theme_scores.items() if vals}
    total = sum(weights.values())
    return {k: v / total for k, v in weights.items()} if total > 0 else {}

def topsis(matrix, weights):
    norm = np.linalg.norm(matrix, axis=0)
    norm_matrix = matrix / norm
    weighted = norm_matrix * weights
    ideal = np.max(weighted, axis=0)
    anti = np.min(weighted, axis=0)
    d_pos = np.linalg.norm(weighted - ideal, axis=1)
    d_neg = np.linalg.norm(weighted - anti, axis=1)
    with np.errstate(divide='ignore', invalid='ignore'):
        closeness = d_neg / (d_pos + d_neg)
        closeness = np.nan_to_num(closeness, nan=0.0, posinf=0.0, neginf=0.0)


    # ✅ Fix NaN or infinite values
    closeness = np.nan_to_num(closeness, nan=0.0, posinf=0.0, neginf=0.0)
    return closeness.round(4)

# === Streamlit UI ===
st.markdown("""
<h1 style='font-size: 40px;'>
    <span style='color:#00274C; font-weight: bold; font-family: cursive;'>Blue</span>
    <span style='color:#FFCB05; font-weight: bold; font-family: cursive;'>Choice </span>
    <span style='font-weight: normal;'>: A Community Centric WEC Decision Support Tool</span>
</h1>
""", unsafe_allow_html=True)
st.markdown(
    """
    <style>
    .stTabs [data-baseweb="tab"] {
        font-weight: 600;
        padding: 10px 16px;
    }
    </style>
    """,
    unsafe_allow_html=True
)
tabs = st.tabs(["1️⃣ Expert Input", "2️⃣ Community Feedback", "3️⃣ Final Ranking"])

# === TAB 1: EXPERT INPUT ===
with tabs[0]:
    st.warning("*All fields are mandatory")
    sheet = connect_to_google_sheets()
    create_themes_tab_with_subcriteria(sheet)
    create_sheet_if_missing(sheet, "Expert Tab")
    ensure_headers_if_missing("Expert Tab")
    create_sheet_if_missing(sheet, "Expert Contributions Tab")
    ensure_headers_if_missing("Expert Contributions Tab")

    expert_scores = {}
    expert_pcms = {}
    inconsistent_themes = []
    expertise_levels = {}

    st.subheader("Expert Scoring: Expertise Level")
    col1, spacer, col2 = st.columns([3, 0.3, 2])

    with col1:
        expert_name = st.text_input("👤 Your Name", placeholder="Enter your name here")

        if expert_name.strip() == "":
            st.warning("Please enter your name before continuing.")
        else:
            st.markdown("### 📊 Self-Rated Expertise")
            for theme in THEMES:
                expertise_levels[theme] = st.slider(
                    f"Rate your expertise in scoring **{theme}** (1 = low, 5 = high)", 1, 5, 3
                )


    with col2:
        st.markdown("### 📘 Theme Descriptions")
        theme_descriptions = {
            "Functional Efficiency": "Ability of the technology to provide reliable and affordable electricity to the community.",
            "Environmental Sustainability": "Minimization of negative impacts on marine ecosystems, wildlife, and human health.",
            "Sense of Place": "Respect for the cultural, visual, and emotional identity of the community and natural surroundings.",
            "Community Prosperity": "Opportunities for local economic development, job creation, tourism, and indirect financial benefits.",
            "Marine Space Utilization": "Efficient and harmonious coexistence with existing maritime activities such as fishing and recreation."
        }

        table_html = """
        <style>
            table {
                width: 100%;
                border-collapse: collapse;
                font-size: 14px;
            }
            th, td {
                border: 1px solid #DDD;
                padding: 8px;
                text-align: left;
            }
            th {
                background-color: #00274C;
                color: #FFCB05;
            }
        </style>
        <table>
            <thead>
                <tr><th>Theme</th><th>Description</th></tr>
            </thead>
            <tbody>
        """

        for theme, desc in theme_descriptions.items():
            table_html += f"<tr><td>{theme}</td><td>{desc}</td></tr>"

        table_html += "</tbody></table>"
        st.markdown(table_html, unsafe_allow_html=True)

    st.subheader("Expert Scoring: Pairwise Comparisons of WEC Designs")
    st.info("Scroll to the bottom to see the Pairwise Comparison Slider Guide")

    st.markdown("""
    <div style="padding: 10px; border: 1px solid #ddd; background-color: #f9f9f9; font-size: 15px">
        All criteria and subcriteria are treated as <strong>benefit criteria</strong>. 
        <br>That means <span style="color: green;"><strong>higher values are always better</strong></span>.
        No inversions or cost adjustments are applied.
    </div>
    """, unsafe_allow_html=True)

    for theme in THEMES:
        st.markdown(
            f"<h4 style='color:#00274C; background-color:#FFCB05; padding:6px'>{theme}</h4>",
            unsafe_allow_html=True,
        )
        for sub_idx, subcriterion in enumerate(THEME_SUBCRITERIA[theme], 1):
            st.markdown(
                f"<h5 style='font-weight:600; margin-left:10px;'>{sub_idx}. {subcriterion}</h5>",
                unsafe_allow_html=True,
            )

            pcm = np.ones((len(WEC_DESIGNS), len(WEC_DESIGNS)))
            for i in range(len(WEC_DESIGNS)):
                for j in range(i + 1, len(WEC_DESIGNS)):
                    st.caption(f"⬅️ Slide left if **{WEC_DESIGNS[j]}** is better, ➡️ slide right if **{WEC_DESIGNS[i]}** is better.")

                    comparison_question = f"""
                        <h3 style='margin-top: 10px; margin-bottom: 6px; font-size: 18px; color:#00274C;'>
                        🔍 In terms of <span style="color:#FFCB05;"><strong>{subcriterion}</strong></span>, how much better is <span style="color:#1f77b4;"><strong>{WEC_DESIGNS[i]}</strong></span> than 
                        <span style="color:#d62728;"><strong>{WEC_DESIGNS[j]}</strong></span>?
                        </h3>
                        """
                    st.markdown(comparison_question, unsafe_allow_html=True)
                    st.markdown("""
                    <div style='
                        display: flex;
                        justify-content: space-between;
                        background-color: #f9f9f9;
                        padding: 6px 0;
                        margin-bottom: 10px;
                        border: 1px solid #ccc;
                        border-radius: 6px;
                        font-size: 15px;
                        font-weight: 500;
                        color: #333;
                    '>
                        <div>-9</div>
                        <div>-7</div>
                        <div>-5</div>
                        <div>-3</div>
                        <div>1</div>
                        <div>3</div>
                        <div>5</div>
                        <div>7</div>
                        <div>9</div>
                    </div>
                    """, unsafe_allow_html=True)

                    scale_options = [-9, -7, -5, -3, 1, 3, 5, 7, 9]

                    descriptors = {
                        -9: f"{WEC_DESIGNS[j]} ≫ {WEC_DESIGNS[i]}",
                        -7: f"{WEC_DESIGNS[j]} > {WEC_DESIGNS[i]}",
                        -5: f"{WEC_DESIGNS[j]} > {WEC_DESIGNS[i]}",
                        -3: f"{WEC_DESIGNS[j]} > {WEC_DESIGNS[i]}",
                        1: f"{WEC_DESIGNS[i]} = {WEC_DESIGNS[j]}",
                        3: f"{WEC_DESIGNS[i]} > {WEC_DESIGNS[j]}",
                        5: f"{WEC_DESIGNS[i]} > {WEC_DESIGNS[j]}",
                        7: f"{WEC_DESIGNS[i]} > {WEC_DESIGNS[j]}",
                        9: f"{WEC_DESIGNS[i]} ≫ {WEC_DESIGNS[j]}"
                    }

                    val = st.select_slider(
                        label="Pairwise Comparison Slider",
                        options=scale_options,
                        value=1,
                        format_func=lambda x: f"{abs(x)} ({descriptors[x]})",
                        key=f"{theme}_{subcriterion}_{i}_{j}",
                        label_visibility="collapsed"
                    )

                    

                    if val == 1:
                        pcm[i, j] = 1
                        pcm[j, i] = 1
                    elif val > 1:
                        pcm[i, j] = val
                        pcm[j, i] = 1 / val
                    elif val < 1:
                        pcm[i, j] = 1 / abs(val)
                        pcm[j, i] = abs(val)

            cr, lam_max, ci = calculate_consistency_index(pcm)
            st.markdown(
                f"<div style='margin-left:10px;'>λₘₐₓ = {lam_max:.3f}, CI = {ci:.3f}, <strong>CR = {cr:.3f}</strong></div>",
                unsafe_allow_html=True,
            )

            if cr > 0.1:
                st.warning(
                    f"⚠️ Subcriterion '{subcriterion}' under theme '{theme}' is inconsistent (CR > 0.1). Please revise."
                )
                inconsistent_themes.append(f"{theme} - {subcriterion}")
            else:
                # ✅ Save weights
                expert_scores[(theme, subcriterion)] = ahp_weights(pcm)
                expert_pcms[(theme, subcriterion)] = pcm

                # ✅ Show PCM as dataframe
                df_pcm = pd.DataFrame(pcm, index=WEC_DESIGNS, columns=WEC_DESIGNS)
                st.markdown(f"**Pairwise Comparison Matrix for:** _{subcriterion}_")
                st.dataframe(df_pcm.style.format("{:.3f}"))

                # ✅ Show Heatmap
                # fig, ax = plt.subplots(figsize=(4.5, 3.5))
                # sns.heatmap(df_pcm, annot=True, fmt=".2f", cmap="YlGnBu", cbar=True, ax=ax)
                # ax.set_title(f"Heatmap: {subcriterion}", fontsize=10)
                # ax.tick_params(axis='x', labelrotation=45)
                # st.pyplot(fig)


            expert_scores[(theme, subcriterion)] = ahp_weights(pcm)
            expert_pcms[(theme, subcriterion)] = pcm

    if not inconsistent_themes:
        if expert_name.strip() == "":
            st.warning("Please enter your name before saving expert scores.")
        elif st.button("✅ Save Expert Scores"):
            # Proceed with saving logic

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            sheet_name = "Expert Contributions Tab"

            # Define headers
            header = ["Timestamp", "Name", "Theme", "Subcriterion", "Expertise Level"]
            for i in range(len(WEC_DESIGNS)):
                for j in range(i + 1, len(WEC_DESIGNS)):
                    header.append(f"PCM_{WEC_DESIGNS[i]}_vs_{WEC_DESIGNS[j]}")

            # Construct new rows
            new_rows = []
            for (theme, subcriterion), pcm in expert_pcms.items():
                flat_pcm = []
                for i in range(len(WEC_DESIGNS)):
                    for j in range(i + 1, len(WEC_DESIGNS)):
                        flat_pcm.append(pcm[i, j])
                level = expertise_levels[theme]
                row = [timestamp, expert_name, theme, subcriterion, level] + flat_pcm
                new_rows.append(row)

            new_df = pd.DataFrame(new_rows, columns=header)

            try:
                # Try to load existing data
                existing_df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
                combined_df = pd.concat([existing_df, new_df], ignore_index=True)
            except FileNotFoundError:
                # File doesn't exist; create new
                combined_df = new_df
            except ValueError:
                # Sheet doesn't exist; treat as new
                combined_df = new_df

            # Write back to Excel
            with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                combined_df.to_excel(writer, sheet_name=sheet_name, index=False)

            st.success("✅ Expert input saved successfully!")

        # 🧮 Button to generate AHP Decision Matrix from Expert Contributions
        with st.expander("🔧 Optional: Generate AHP Decision Matrix"):
            st.markdown(
                "This computes the **expertise-weighted average AHP weights** across all expert inputs for each subcriterion "
                "and saves them as a decision matrix in Excel."
            )
            
            if st.button("📥 Generate AHP Decision Matrix"):
                try:
                    df = pd.read_excel(EXCEL_FILE, sheet_name="Expert Contributions Tab")

                    # Prepare AHP Decision Matrix: rows = WECs, columns = subcriteria
                    decision_matrix = pd.DataFrame(index=WEC_DESIGNS)

                    grouped = df.groupby(["Theme", "Subcriterion"])

                    for (theme, subcriterion), group in grouped:
                        weighted_sums = np.zeros(len(WEC_DESIGNS))
                        total_weight = 0

                        for _, row in group.iterrows():
                            try:
                                # Get expertise level for this theme and normalize it (scale 1–5)
                                expertise = float(row["Expertise Level"])
                                weight = expertise / 5.0
                                
                                # Build PCM matrix from the row
                                pcm_values = row[5:].values.astype(float)
                                pcm = np.ones((len(WEC_DESIGNS), len(WEC_DESIGNS)))

                                idx = 0
                                for i in range(len(WEC_DESIGNS)):
                                    for j in range(i + 1, len(WEC_DESIGNS)):
                                        pcm[i, j] = pcm_values[idx]
                                        pcm[j, i] = 1 / pcm_values[idx]
                                        idx += 1

                                weights = ahp_weights(pcm)  # AHP weights
                                weighted_sums += weight * weights
                                total_weight += weight
                            except Exception as e:
                                st.warning(f"⚠️ Skipped a row due to error: {e}")

                        if total_weight > 0:
                            avg_weights = weighted_sums / total_weight
                            decision_matrix[f"{theme} - {subcriterion}"] = avg_weights
                        else:
                            st.warning(f"⚠️ No valid expert inputs found for: {theme} - {subcriterion}")

                    decision_matrix = decision_matrix.fillna(0)

                    # Save to Excel
                    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                        decision_matrix.to_excel(writer, sheet_name="AHP Decision Matrix")

                    st.success("✅ AHP Decision Matrix saved to Excel successfully!")
                    st.markdown("### 🧮 AHP Decision Matrix")
                    st.dataframe(decision_matrix.style.format("{:.3f}"))

                except Exception as e:
                    st.error(f"⚠️ Error generating AHP Decision Matrix: {e}")

    else:
        st.error("❌ One or more subcriteria are inconsistent. Fix them before saving.")


    st.markdown("""
    <style>
        .comparison-guide {
            font-size: 14px;
            margin-top: 30px;
        }
        .comparison-guide table {
            border-collapse: collapse;
            width: 100%;
        }
        .comparison-guide th, .comparison-guide td {
            border: 1px solid #ddd;
            padding: 8px;
            text-align: center;
        }
        .comparison-guide th {
            background-color: #003366;
            color: #FFCB05;
        }
        .comparison-guide caption {
            caption-side: top;
            text-align: left;
            font-weight: bold;
            font-size: 18px;
            margin-bottom: 10px;
        }
        .example {
            background-color: #f1f8ff;
            border-left: 4px solid #1f77b4;
            padding: 10px;
            margin-top: 20px;
            font-size: 14px;
        }
    </style>

    <div class="comparison-guide">
    <h4>Pairwise Comparison Slider Guide</h4>
    <table>
        <tr>
            <th>Slider Value</th>
            <th>Meaning</th>
        </tr>
        <tr><td><b>1</b></td><td>Both WEC designs perform equally under this subcriterion</td></tr>
        <tr><td><b>3</b></td><td>First design performs slightly better than the second</td></tr>
        <tr><td><b>5</b></td><td>First design clearly performs better than the second</td></tr>
        <tr><td><b>7</b></td><td>First design strongly outperforms the second</td></tr>
        <tr><td><b>9</b></td><td>First design is overwhelmingly better under this subcriterion</td></tr>
        <tr><td><b>2, 4, 6, 8</b></td><td>Intermediate judgments between the above levels</td></tr>
    </table>
    </div>

    <div class="example">
    <b>📌 Example:</b><br>
    If you set <b>Design A vs Design B (Noise) = 5</b>, you’re saying <i>"Design A"</i> is <b>clearly better</b> than <i>"Design B"</i> in minimizing noise impact.
    </div>
    """, unsafe_allow_html=True)


# === TAB 2: COMMUNITY INPUT ===
import json
from datetime import datetime

with tabs[1]:
    st.subheader("🗣️ Community Survey: Views on Wave Energy for Beaver Island")

    st.markdown("## Q1. How long have you lived on Beaver Island?")
    q1 = st.radio("", ["Less than 1 year", "1–5 years", "6–10 years", "11–20 years", "More than 20 years"], key="q1")

    st.markdown("## Q2. How familiar are you with wave energy extraction technologies?")
    q2 = st.radio("", ["Not familiar at all", "Slightly familiar", "Moderately familiar", "Very familiar", "I have seen or interacted with such systems"], key="q2")

    st.markdown("## Q3. Concern about reliability of the energy supply")
    q3 = st.radio("How concerned are you about the reliability of the energy supply on Beaver Island?", ["Not at all concerned", "Slightly concerned", "Moderately concerned", "Concerned", "Very concerned"], key="q3")
    q3_reason = st.text_area(
        "Q3(b) Please explain why you are not concerned about reliability of the energy supply on Beaver Island:" if q3 == "Not at all concerned" else
        "Q3(a) Please describe your concerns about reliability of the energy supply on Beaver Island:",
        key="q3_reason"
    )

    st.markdown("## Q4. Support for wave energy development")
    q4 = st.radio("Would you support the development of wave energy technology to enhance Beaver Island’s energy security over the next five years?",
        ["Strongly oppose", "Somewhat oppose", "Neither support nor oppose", "Somewhat support", "Strongly support"], key="q4")
    if q4 in ["Strongly oppose", "Somewhat oppose"]:
        q4_reason = st.text_area("Q4(a) Please explain why do you oppose development of wave energy technology to enhance Beaver Island’s energy security over the next five years:", key="q4_reason")
    elif q4 in ["Somewhat support", "Strongly support"]:
        q4_reason = st.text_area("Q4(b) Please explain why do you support development of wave energy technology to enhance Beaver Island’s energy security over the next five years:", key="q4_reason")
    else:
        q4_reason = st.text_area("Q4(c) Please explain why you're neutral towards development of wave energy technology to enhance Beaver Island’s energy security over the next five years:", key="q4_reason")

    st.markdown("## Q5. Do you think wave energy would be beneficial for Beaver Island?")
    q5 = st.radio("", ["Yes", "No"], key="q5")
    q5_reason = st.text_area("Q5(a) In what ways do you think wave energy could benefit Beaver Island:" if q5 == "Yes" else "Q5(b) Please explain why you think wave energy may not be beneficial for Beaver Island:", key="q5_reason")

    def ask_concern_question(label, not_concerned_prompt, concerned_prompt, qkey):
        response = st.radio(label, ["Not at all concerned", "Slightly concerned", "Moderately concerned", "Concerned", "Very concerned"], key=qkey)
        reason = st.text_area(not_concerned_prompt if response == "Not at all concerned" else concerned_prompt, key=f"{qkey}_reason")
        return response, reason

    st.markdown("## Q6. Concern about affordability")
    q6, q6_reason = ask_concern_question(
        "How concerned are you that wave energy may not improve the affordability of electricity on Beaver Island?",
        "Q6(a) Please tell us why you feel confident that wave energy on Beaver Island will be an affordable solution for meeting local energy needs.",
        "Q6(b) What aspects of affordability concern you most regarding wave energy on Beaver Island?",
        qkey="q6"
    )

    st.markdown("## Q7. Concern about marine ecosystems or wildlife")
    q7, q7_reason = ask_concern_question(
        "How concerned are you that wave energy development on Beaver Island could negatively affect marine ecosystems or wildlife?",
        "Q7(a) What makes you confident that wave energy on Beaver Island will be environmentally safe?",
        "Q7(b) What specific environmental impacts are you most concerned about from wave energy on Beaver Island?",
        qkey="q7"
    )

    st.markdown("## Q8. Concern about natural views and shoreline aesthetics")
    q8, q8_reason = ask_concern_question(
        "How concerned are you that wave energy devices on Beaver Island may negatively impact natural views, the shoreline, or the island’s appearance?",
        "Q8(a) Why do you believe wave energy devices on Beaver Island will not affect the island’s natural appearance?",
        "Q8(b) What concerns do you have about the placement or visual impact of wave energy devices on Beaver Island?",
        qkey="q8"
    )

    st.markdown("## Q9. Concern about economic benefits")
    q9, q9_reason = ask_concern_question(
        "How concerned are you that wave energy development on Beaver Island may not provide meaningful benefits to the local economy?",
        "Q9(a) What gives you confidence that wave energy will support Beaver Island’s economy?",
        "Q9(b) What concerns do you have about the economic impacts of wave energy development on Beaver Island?",
        qkey="q9"
    )

    st.markdown("## Q10. Would you like to be contacted in future to help validate our results?")
    q10 = st.radio("Would you like to be contacted in future to help validate our results?", ["Yes", "No"], key="q10")
    contact_name = contact_email = contact_phone = ""
    if q10 == "Yes":
        contact_name = st.text_input("Full Name", key="contact_name")
        contact_email = st.text_input("Email", key="contact_email")
        contact_phone = st.text_input("Phone", key="contact_phone")

    if st.button("✉️ Submit Survey Response", key="submit_button"):
        likert_map_5pt = {
            "Not at all concerned": 1, "Slightly concerned": 2, "Moderately concerned": 3, "Concerned": 4, "Very concerned": 5,
            "Not familiar at all": 1, "Slightly familiar": 2, "Moderately familiar": 3, "Very familiar": 4, "I have seen or interacted with such systems": 5,
            "Strongly oppose": 1, "Somewhat oppose": 2, "Neither support nor oppose": 3, "Somewhat support": 4, "Strongly support": 5,
        }
        likert_map_2pt = {"Yes": 1, "No": 0}

        # WHFS Scoring
        whfs_q3 = get_whfs_from_likert_and_text(likert_map_5pt.get(q3, ""), q3_reason)
        whfs_q4 = get_whfs_from_likert_and_text(likert_map_5pt.get(q4, ""), q4_reason)
        whfs_q5 = get_whfs_from_likert_and_text(likert_map_2pt.get(q5, ""), q5_reason)
        whfs_q6 = get_whfs_from_likert_and_text(likert_map_5pt.get(q6, ""), q6_reason)
        whfs_q7 = get_whfs_from_likert_and_text(likert_map_5pt.get(q7, ""), q7_reason)
        whfs_q8 = get_whfs_from_likert_and_text(likert_map_5pt.get(q8, ""), q8_reason)
        whfs_q9 = get_whfs_from_likert_and_text(likert_map_5pt.get(q9, ""), q9_reason)

        row = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            q1,
            likert_map_5pt.get(q2, ""),
            likert_map_5pt.get(q3, ""), q3_reason, json.dumps(whfs_q3),
            likert_map_5pt.get(q4, ""), q4_reason, json.dumps(whfs_q4),
            likert_map_2pt.get(q5, ""), q5_reason, json.dumps(whfs_q5),
            likert_map_5pt.get(q6, ""), q6_reason, json.dumps(whfs_q6),
            likert_map_5pt.get(q7, ""), q7_reason, json.dumps(whfs_q7),
            likert_map_5pt.get(q8, ""), q8_reason, json.dumps(whfs_q8),
            likert_map_5pt.get(q9, ""), q9_reason, json.dumps(whfs_q9),
            q10,
            contact_name, contact_email, contact_phone
        ]

        headers = [
            "Timestamp", "Q1 Residence Duration", "Q2 Familiarity (1-5)",
            "Q3 Reliability Concern", "Q3 Explanation", "Q3 WHFS",
            "Q4 Support Level", "Q4 Explanation", "Q4 WHFS",
            "Q5 Benefit Belief", "Q5 Explanation", "Q5 WHFS",
            "Q6 Affordability Concern", "Q6 Explanation", "Q6 WHFS",
            "Q7 Ecosystem Concern", "Q7 Explanation", "Q7 WHFS",
            "Q8 Visual Concern", "Q8 Explanation", "Q8 WHFS",
            "Q9 Economic Concern", "Q9 Explanation", "Q9 WHFS",
            "Q10 Contact?", "Name", "Email", "Phone"
        ]

        try:
            ws, wb = read_excel("Community Feedback Tab")
            if ws.max_row == 0 or all(cell.value is None for cell in ws[1]):
                ws.append(headers)
            ws.append(row)
            wb.save(EXCEL_FILE)
            st.success("✅ Your survey response has been saved.")
        except Exception as e:
            st.error(f"⚠️ Failed to save response: {e}")



# === TAB 3: FINAL RANKING ===
with tabs[2]:
    st.subheader("Final WEC Ranking using Combined AHP + WHFS-TOPSIS")

    create_sheet_if_missing(EXCEL_FILE, "Final Rankings")
    create_sheet_if_missing(EXCEL_FILE, "Final AHP Theme Weights")
    ensure_headers_if_missing("Final Rankings")

    if st.button("🏁 Run Ranking"):
        try:
            # --- Load AHP Decision Matrix ---
            ahp_df = pd.read_excel(EXCEL_FILE, sheet_name="AHP Decision Matrix", index_col=0)

            # --- Load Preliminary Survey ---
            prelim_df = pd.read_excel("BI_Preliminary_Survey.xlsx")
            prelim_df = prelim_df.dropna(subset=["Theme Percentage %"])
            prelim_theme_weights = prelim_df.groupby("Theme")["Theme Percentage %"].mean()
            prelim_theme_weights /= prelim_theme_weights.sum()

            # --- Load Community WHFS Data ---
            comm_ws, _ = read_excel("Community Feedback Tab")
            comm_df = pd.DataFrame(comm_ws.values)
            comm_df.columns = comm_df.iloc[0]
            comm_df = comm_df[1:]

            whfs_question_map = {
                "Q3 WHFS": "Functional Efficiency",
                "Q6 WHFS": "Functional Efficiency",
                "Q7 WHFS": "Environmental Sustainability",
                "Q8 WHFS": "Sense of Place",
                "Q9 WHFS": "Community Prosperity"
            }

            from ast import literal_eval
            whfs_theme_agg = {t: [] for t in set(whfs_question_map.values())}
            for _, row in comm_df.iterrows():
                for q, theme in whfs_question_map.items():
                    try:
                        scores = literal_eval(row[q]) if isinstance(row[q], str) else row[q]
                        for level, wt in scores:
                            whfs_theme_agg[theme].append(level * wt)
                    except:
                        continue

            whfs_theme_weights = {t: np.mean(v) for t, v in whfs_theme_agg.items() if v}
            whfs_total = sum(whfs_theme_weights.values())
            whfs_theme_weights = {k: v / whfs_total for k, v in whfs_theme_weights.items()} if whfs_total > 0 else {}

            # --- Combine Preliminary and WHFS Theme Weights ---
            all_themes = sorted(set(prelim_theme_weights.index).union(whfs_theme_weights.keys()))
            combined_theme_weights = {
                t: 0.5 * prelim_theme_weights.get(t, 0) + 0.5 * whfs_theme_weights.get(t, 0)
                for t in all_themes
            }

            # --- Group Subcriteria by Theme and Apply Theme Weights ---
            grouped_scores = {}
            for theme in combined_theme_weights:
                sub_cols = [c for c in ahp_df.columns if c.startswith(theme)]
                if sub_cols:
                    theme_score = ahp_df[sub_cols].mean(axis=1)
                    grouped_scores[theme] = theme_score * combined_theme_weights[theme]

            final_matrix = pd.DataFrame(grouped_scores)
            
            # Normalize each row (WEC Design) so that the sum of theme scores equals 1
            normalized_matrix = final_matrix.copy()
            normalized_matrix = normalized_matrix.drop(columns=["Total Score"], errors='ignore')
            row_sums = normalized_matrix.sum(axis=1)
            normalized_matrix = normalized_matrix.div(row_sums, axis=0)
            normalized_matrix["Total Score"] = 1.0  # By construction, every row now sums to 1

            # --- Save Final Weighted AHP Scores ---
            with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                normalized_matrix.to_excel(writer, sheet_name="Final AHP Theme Weights")

            # --- TOPSIS Ranking ---
            matrix = normalized_matrix[all_themes].values
            weights = np.array([combined_theme_weights[t] for t in all_themes])
            closeness = topsis(matrix, weights)

            result_df = pd.DataFrame({
                "WEC Design": final_matrix.index,
                "Closeness to Ideal": closeness
            }).sort_values(by="Closeness to Ideal", ascending=False)
            st.session_state["result_df"] = result_df

            # --- Save Final Rankings ---
            ws, wb = read_excel("Final Rankings")
            ws.delete_rows(1, ws.max_row)
            ws.append(["WEC Design", "Closeness to Ideal"])
            for row in result_df.values.tolist():
                ws.append(row)
            wb.save(EXCEL_FILE)

            # --- Display Outputs ---
            st.markdown("### ✅ Final AHP Scores (Normalized)")
            st.dataframe(normalized_matrix.drop(columns=["Total Score"]).round(4))

            # === Individual Radar Charts for Each WEC Design ===
            from textwrap import fill
            labels = list(normalized_matrix.columns.drop("Total Score"))
            wrapped_labels = [fill(label, width=15) for label in labels]

            num_vars = len(labels)
            angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()
            angles += angles[:1]

            # Define color map
            cmap = cm.get_cmap('tab10')
            colors = cmap.colors  # a list of distinct colors

            st.markdown("### 🕸️ AHP Radar Charts")

            # Streamlit columns
            cols = st.columns(len(normalized_matrix))  # dynamically create one column per WEC

            for i, (idx, row) in enumerate(normalized_matrix.iterrows()):
                values = row[labels].tolist()
                values += values[:1]

                fig, ax = plt.subplots(figsize=(4, 3), subplot_kw=dict(polar=True))
                ax.plot(angles, values, color=colors[i % len(colors)], linewidth=1.5)
                ax.fill(angles, values, color=colors[i % len(colors)], alpha=0.25)

                ax.set_xticks([])

                for angle in angles[:-1]:
                    ax.plot([angle, angle], [0, 0.5], color='gray', linestyle='--', linewidth=0.6, alpha=0.6)

                for angle, label in zip(angles[:-1], wrapped_labels):
                    ax.text(angle, 0.4, label, ha='center', va='center', fontsize=7, wrap=True)

                ax.set_theta_offset(np.pi / 2)
                ax.set_theta_direction(-1)
                ax.set_ylim(0, 0.3)
                ax.set_rlabel_position(180 / num_vars)
                ax.tick_params(axis='y', labelsize=6)
                ax.set_yticklabels([])
                ax.grid(True, linestyle='--', linewidth=0.5, alpha=0.6)

                # Add WEC title below the plot
                fig.text(0.5, -0.08, f"{idx}", ha='center', fontsize=10)

                safe_name = idx.replace(" ", "_").replace("/", "_").lower()
                image_path = os.path.join(os.path.dirname(EXCEL_FILE), f"{safe_name}_radar_chart.png")
                fig.savefig(image_path, dpi=300, bbox_inches='tight')

                cols[i].pyplot(fig)

            col1, col2 = st.columns([1, 1])  # Two equal columns

            with col1:
                labels = [fill(label, width=15) for label in combined_theme_weights.keys()]
                values = list(combined_theme_weights.values())

                fig, ax = plt.subplots(figsize=(2.5, 2))
                bars = ax.bar(labels, values)

                ax.set_title("Combined Theme Weights from Surveys", fontsize=5)
                ax.set_ylabel("Weight", fontsize=4)

                plt.xticks(rotation=45, ha='right', fontsize=4)
                ax.tick_params(axis='y', labelsize=4)

                for bar, val in zip(bars, values):
                    ax.text(bar.get_x() + bar.get_width() / 2, val + 0.01, f"{val:.2f}", 
                            ha='center', va='bottom', fontsize=4)

                ax.margins(y=0.2)
                plt.tight_layout()
                st.pyplot(fig)

            with col2:
                # Add vertical space to roughly center-align with col1 chart
                st.markdown("<br><br><br><br><br><br>", unsafe_allow_html=True)  # adjust number of <br> as needed
                st.markdown("#### 🏆 TOPSIS Final Rankings")
                st.dataframe(result_df)
            
            st.session_state["normalized_matrix"] = normalized_matrix
            st.session_state["combined_theme_weights"] = combined_theme_weights
            st.session_state["all_themes"] = all_themes
            st.session_state["wec_names"] = list(normalized_matrix.index)
        except Exception as e:
            st.error(f"❌ Error in final ranking: {e}")
    
    # --- Monte Carlo Sensitivity Analysis ---
    st.markdown("### 🎲 Monte Carlo Sensitivity Analysis (Theme Weight Uncertainty)")

    if "normalized_matrix" in st.session_state:
        if st.button("▶️ Run 10,000 Simulations"):
            try:
                K = 10000
                normalized_matrix = st.session_state["normalized_matrix"]
                combined_theme_weights = st.session_state["combined_theme_weights"]
                all_themes = st.session_state["all_themes"]
                wec_names = st.session_state["wec_names"]

                # Run Monte Carlo as before...
                decision_matrix = normalized_matrix[all_themes].values
                alpha = np.array([combined_theme_weights[t] for t in all_themes])
                alpha_scaled = alpha * 100

                rank_distributions = []
                for _ in range(K):
                    sampled_weights = np.random.dirichlet(alpha_scaled)
                    scores = topsis(decision_matrix, sampled_weights)
                    ranks = pd.Series((-scores).argsort().argsort() + 1, index=wec_names)
                    rank_distributions.append(ranks)

                rank_df = pd.DataFrame(rank_distributions)

                col1, col2 = st.columns([1.5, 1])  # Adjust width ratio as needed

                with col1:
                    st.markdown("#### 📦 Rank Distribution Across Simulations")
                    fig, ax = plt.subplots(figsize=(6, 3.5))
                    sns.boxplot(data=rank_df[wec_names], ax=ax)
                    sampled_df = rank_df.sample(n=500, random_state=42)
                    # Overlay stripplot with full 10,000 samples
                    sns.stripplot(
                        data=rank_df[wec_names],
                        ax=ax,
                        color=".25",
                        size=1.5,
                        jitter=0.25,
                        alpha=0.3
                    )
                    ax.set_ylabel("Rank", fontsize=9)
                    ax.set_title("Rank Stability Across 10,000 Monte Carlo Runs", fontsize=10)
                    ax.tick_params(axis='x', labelrotation=0, labelsize=7)
                    ax.tick_params(axis='y', labelsize=7)
                    ax.set_ylim(0.5, len(wec_names) + 1)
                    plt.tight_layout()
                    st.pyplot(fig)

                with col2:
                    st.markdown("<br><br><br><br><br><br>", unsafe_allow_html=True)  # adjust number of <br> as needed
                    st.markdown("#### 📊 Summary of Rank Stability")
                    stats_df = pd.DataFrame({
                        "WEC Design": wec_names,
                        "Mean Rank": rank_df.mean().values,
                        "Std Dev": rank_df.std().values
                    }).sort_values("Mean Rank")
                    st.dataframe(stats_df.round(2), use_container_width=True)


                from scipy.stats import kendalltau

                # Get the baseline ranking from result_df
                baseline_ranking = st.session_state["result_df"]["WEC Design"].tolist()
                baseline_order = {name: rank for rank, name in enumerate(baseline_ranking)}

                # Compute Kendall's Tau distance for each simulation
                tau_distances = []
                for k in range(rank_df.shape[0]):
                    sim_ranking = rank_df.iloc[k].sort_values().index.tolist()
                    sim_order = [baseline_order[name] for name in sim_ranking]
                    tau, _ = kendalltau(list(range(len(sim_order))), sim_order)
                    tau_distances.append(1 - tau)  # distance = 1 - similarity

                col1, col2 = st.columns([1.5, 1])  # Wider plot, narrower stats

                with col1:
                    st.markdown("#### 📉 Kendall’s Tau Distance from Baseline")
                    fig, ax = plt.subplots(figsize=(5.5, 3.2))
                    ax.hist(tau_distances, bins=30, color='slateblue', edgecolor='black', alpha=0.7)
                    ax.set_xlabel("Kendall’s Tau Distance", fontsize=9)
                    ax.set_ylabel("Frequency", fontsize=9)
                    ax.set_title("Distribution of Kendall’s Tau Distance over 10,000 Simulations", fontsize=10)
                    ax.tick_params(axis='both', labelsize=7)
                    plt.tight_layout()
                    st.pyplot(fig)

                with col2:
                    st.markdown("<br><br><br><br><br><br>", unsafe_allow_html=True)  # adjust number of <br> as needed
                    mean_tau = np.mean(tau_distances)
                    std_tau = np.std(tau_distances)
                    st.markdown("#### 📈 Kendall’s Tau Summary")
                    st.markdown(f"""
                    - **Mean:** {mean_tau:.5f}  
                    - **Std Dev:** {std_tau:.5f}
                    """)

                # --- Save Monte Carlo Simulation Results ---
                output_dir = os.path.dirname(EXCEL_FILE)
                rank_csv_path = os.path.join(output_dir, "montecarlo_rank_distributions.csv")
                tau_csv_path = os.path.join(output_dir, "montecarlo_kendall_tau.csv")

                rank_df.to_csv(rank_csv_path, index=False)
                pd.DataFrame({"KendallTauDistance": tau_distances}).to_csv(tau_csv_path, index=False)

                st.success(f"✅ Monte Carlo results saved to:\n- `{rank_csv_path}`\n- `{tau_csv_path}`")

            except Exception as e:
                st.error(f"❌ Error during Monte Carlo simulation: {e}")
    else:
        st.info("⚠️ Please run the ranking first to enable sensitivity analysis.")


st.markdown("---")
st.markdown("### ⚠️ Danger Zone")
if st.button("🧹 Clear All Data from Google Sheets"):
    clear_all_google_sheet_tabs()

st.markdown(
    "<hr style='margin-top: 50px;'><div style='text-align: center; font-size: 18px; color: gray;'>© 2025 Vishnu Vijayasankar. All rights reserved.</div>",
    unsafe_allow_html=True
)